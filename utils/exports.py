import io
import os
import tempfile
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fpdf import FPDF

MESES_PDF = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
             7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}


def _latin1(s: str) -> str:
    return str(s).encode('latin-1', errors='replace').decode('latin-1')


def df_to_excel_bytes(df: pd.DataFrame, titulo: str, filtro: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    hdr_fill = PatternFill('solid', fgColor='003087')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    alt_fill = PatternFill('solid', fgColor='EEF2FF')

    ws['A1'] = f'DIRESA HUANCAVELICA — {titulo}'
    ws['A1'].font = Font(bold=True, color='003087', size=13)
    # MINOR-6: derivar periodo de los datos reales
    if not df.empty and 'mes' in df.columns:
        meses_datos = sorted(df['mes'].dropna().unique().astype(int))
        if meses_datos:
            m_ini = MESES_PDF.get(meses_datos[0], str(meses_datos[0])).upper()
            m_fin = MESES_PDF.get(meses_datos[-1], str(meses_datos[-1])).upper()
            periodo_excel = f'{m_ini} - {m_fin} 2026' if m_ini != m_fin else f'{m_ini} 2026'
        else:
            periodo_excel = '2026'
    else:
        periodo_excel = '2026'
    ws['A2'] = f'Filtro: {filtro}  |  PERIODO: {periodo_excel}'
    ws['A2'].font = Font(italic=True, color='444444', size=10)

    max_col = max(len(df.columns), 1)   # MINOR-7: usar ancho real del df
    ws.merge_cells(f'A1:{get_column_letter(max_col)}1')
    ws.merge_cells(f'A2:{get_column_letter(max_col)}2')
    ws.row_dimensions[3].height = 5

    if df.empty:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    RENAME = {
        'año': 'AÑO', 'mes': 'MES', 'provincia': 'PROVINCIA',
        'red': 'RED', 'microred': 'MICRORED', 'eess': 'ESTABLECIMIENTO',
        'renaes': 'RENAES', 'num_doc': 'N° DOCUMENTO', 'nombres': 'NOMBRES',
        'den': 'DENOMINADOR', 'num': 'NUMERADOR', 'pct': '% AVANCE',
    }
    df_out = df.copy()
    # Ordenar por mes si existe
    if 'mes' in df_out.columns:
        df_out = df_out.sort_values('mes')
    if 'pct' in df_out.columns:
        df_out['pct'] = (df_out['pct'] * 100).round(1).astype(str) + '%'
    if 'mes' in df_out.columns:
        df_out['mes'] = df_out['mes'].map(MESES_PDF).fillna(df_out['mes'].astype(str))
    df_out = df_out.rename(columns=RENAME)
    cols = [c for c in df_out.columns if c not in ('ficha_id',)]
    df_out = df_out[cols]

    start = 4
    for ci, col in enumerate(df_out.columns, 1):
        cell = ws.cell(row=start, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for ri, row in enumerate(df_out.itertuples(index=False), start + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal='center')

    for ci in range(1, len(df_out.columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_bytes(df: pd.DataFrame, titulo: str, filtro: str,
                    logro_str: str, periodo: str,
                    map_bytes: bytes | None = None) -> bytes:
    titulo    = _latin1(titulo)
    filtro    = _latin1(filtro.replace('›', '>').replace('→', '>'))
    logro_str = _latin1(logro_str)

    pdf = FPDF()
    pdf.set_margins(6, 6, 6)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=10)

    # ── Encabezado institucional ───────────────────────────────────────────
    pdf.set_fill_color(0, 48, 135)
    pdf.rect(0, 0, 210, 26, 'F')
    pdf.set_fill_color(0, 132, 61)
    pdf.rect(0, 26, 210, 3, 'F')

    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(0, 4)
    pdf.cell(210, 8, 'DIRESA HUANCAVELICA', align='C', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 8)
    pdf.set_x(0)
    pdf.cell(210, 5, f'Indicadores de Desempeno DL 1153-2026  |  {periodo}',
             align='C', new_x='LMARGIN', new_y='NEXT')

    # ── Título del indicador ────────────────────────────────────────────────
    pdf.set_xy(6, 32)
    pdf.set_text_color(0, 48, 135)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.multi_cell(198, 6, titulo, new_x='LMARGIN', new_y='NEXT')

    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(70, 70, 70)
    pdf.set_x(6)
    pdf.cell(99, 5, f'Filtro: {filtro}')
    pdf.cell(99, 5, f'Meta: {logro_str}', align='R', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(2)

    # ── Métricas resumen ────────────────────────────────────────────────────
    den_t = int(df['den'].sum()) if 'den' in df.columns else 0
    num_t = int(df['num'].sum()) if 'num' in df.columns else 0
    pct_t = num_t / den_t * 100 if den_t > 0 else 0
    pend  = max(0, den_t - num_t)
    metrics = [('DENOMINADOR', f'{den_t:,}'), ('NUMERADOR', f'{num_t:,}'),
               ('% AVANCE', f'{pct_t:.1f}%'), ('PENDIENTES', f'{pend:,}')]

    pdf.set_x(6)
    for lbl, _ in metrics:
        pdf.set_fill_color(0, 48, 135)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 6.5)
        pdf.cell(47, 5, lbl, fill=True, align='C', border=0)
    pdf.ln(5)
    pdf.set_x(6)
    for _, val in metrics:
        pdf.set_fill_color(235, 241, 255)
        pdf.set_text_color(0, 48, 135)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.cell(47, 6, val, fill=True, align='C', border=1)
    pdf.ln(9)

    if df.empty:
        pdf.set_text_color(180, 0, 0)
        pdf.cell(198, 8, 'Sin datos para el filtro seleccionado.',
                 new_x='LMARGIN', new_y='NEXT')
        _footer(pdf)
        return bytes(pdf.output())

    y_section = pdf.get_y()

    # ── Mapa PNG (columna izquierda) ────────────────────────────────────────
    if map_bytes:
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                tmp.write(map_bytes)
                tmp_path = tmp.name
            pdf.set_xy(6, y_section)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(0, 48, 135)
            pdf.cell(90, 5, 'Mapa por Red / Provincia', align='C',
                     new_x='LMARGIN', new_y='NEXT')
            pdf.image(tmp_path, x=6, y=y_section + 6, w=90)
        except Exception:
            pass
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)

    # ── Gráfico de barras (columna derecha) ─────────────────────────────────
    col = next((c for c in ['eess', 'microred', 'red']
                if c in df.columns and df[c].str.len().gt(0).any()), None)
    if col:
        agg = (df.groupby(col)
                 .agg(den=('den', 'sum'), num=('num', 'sum'))
                 .reset_index())
        agg['pct'] = np.where(agg['den'] > 0, agg['num'] / agg['den'] * 100, 0)
        agg = agg[agg[col].str.len() > 0].sort_values('pct', ascending=False).head(14)

        try:
            logro_pct = float(logro_str.replace('%', '').strip())
        except Exception:
            logro_pct = 0

        cx = 99
        bar_zone_w = 105
        max_bar = 72
        label_w  = 30
        pdf.set_xy(cx, y_section)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_text_color(0, 48, 135)
        pdf.cell(bar_zone_w, 5, 'Avance por Establecimiento', align='C',
                 new_x='LMARGIN', new_y='NEXT')

        cy = y_section + 7
        for _, row in agg.iterrows():
            pct = min(100.0, float(row['pct']))
            if logro_pct > 0:
                if pct >= logro_pct:           r, g, b = 45, 198, 83
                elif pct >= logro_pct * 0.80:  r, g, b = 255, 183, 3
                else:                          r, g, b = 230, 57, 70
            else:
                r, g, b = 45, 198, 83

            pdf.set_xy(cx, cy)
            pdf.set_font('Helvetica', '', 5.5)
            pdf.set_text_color(50, 50, 50)
            name = _latin1(str(row[col])[:24])
            pdf.cell(label_w, 4, name)

            bar_x = cx + label_w + 1
            bar_w = pct / 100 * max_bar
            pdf.set_fill_color(215, 215, 215)
            pdf.rect(bar_x, cy + 0.5, max_bar, 3.5, 'F')
            pdf.set_fill_color(r, g, b)
            pdf.rect(bar_x, cy + 0.5, bar_w, 3.5, 'F')
            pdf.set_font('Helvetica', 'B', 5)
            pdf.set_text_color(50, 50, 50)
            pdf.set_xy(bar_x + max_bar + 1, cy + 0.2)
            pdf.cell(10, 4, f'{pct:.0f}%')
            cy += 7
            if cy > 270:
                break

    # ── Tabla de Avance Detallada ──────────────────────────────────────────
    table_y = max(y_section + 105, pdf.get_y() + 4)
    if table_y > 250:
        pdf.add_page()
        table_y = 10

    pdf.set_xy(6, table_y)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_text_color(0, 48, 135)
    pdf.cell(0, 5, 'Tabla de Avance Detallada', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(1)

    df_sorted = df.copy()
    sort_cols = [c for c in ['mes', 'red', 'eess'] if c in df_sorted.columns]
    if sort_cols:
        df_sorted = df_sorted.sort_values(sort_cols)

    TCOLS = [c for c in ['mes', 'red', 'eess', 'den', 'num', 'pct'] if c in df_sorted.columns]
    THDRS  = {'mes': 'MES', 'red': 'RED', 'eess': 'EESS',
              'den': 'DEN', 'num': 'NUM', 'pct': '% AVZ'}
    TWIDTHS = {'mes': 16, 'red': 30, 'eess': 50, 'den': 18, 'num': 18, 'pct': 20}

    pdf.set_x(6)
    pdf.set_fill_color(0, 48, 135)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 7)
    for c in TCOLS:
        pdf.cell(TWIDTHS[c], 6, THDRS[c], border=1, fill=True, align='C')
    pdf.ln()

    alt = False
    pdf.set_font('Helvetica', '', 6.5)
    for _, row in df_sorted.iterrows():
        if pdf.get_y() > 282:
            _table_header(pdf, TCOLS, THDRS, TWIDTHS)
        pdf.set_fill_color(238, 242, 255) if alt else pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(30, 30, 30)
        pdf.set_x(6)
        for c in TCOLS:
            v = row[c]
            if c == 'mes':
                v = MESES_PDF.get(int(v), str(v)) if str(v).isdigit() else str(v)
            elif c == 'pct':
                v = f'{float(v)*100:.1f}%'
            else:
                v = str(v)[:22]
            pdf.cell(TWIDTHS[c], 5.5, _latin1(v), border=1, fill=True, align='C')
        pdf.ln()
        alt = not alt

    _footer(pdf)
    return bytes(pdf.output())


def _table_header(pdf, TCOLS, THDRS, TWIDTHS):
    pdf.add_page()
    pdf.set_x(6)
    pdf.set_fill_color(0, 48, 135)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 7)
    for c in TCOLS:
        pdf.cell(TWIDTHS[c], 6, THDRS[c], border=1, fill=True, align='C')
    pdf.ln()
    pdf.set_font('Helvetica', '', 6.5)


def _footer(pdf):
    pdf.set_y(-12)
    pdf.set_font('Helvetica', 'I', 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, f'DIRESA Huancavelica - Sistema DL 1153-2026  |  Pagina {pdf.page_no()}',
             align='C')
